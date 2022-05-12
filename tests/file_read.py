from openpyxl import Workbook, load_workbook
#local imports
from .models import QuestionModel, AnswerModel



def read_new_tests(file, request):
    wb = load_workbook(file)
    ws = wb.active
    letters, blank_limit = 'CDEF', 4
    # blank_limit  ==> is used for checking question values stopped. E.g. I have saved 50 questions in xlsx.
    # and if the next 4 cell is Null, loop will be automatically broken. And be returned response. 
    for i in range(3, 475):
        try:
            question = ws["B{}".format(i)].value
            if question:
                # try:
                #     question_obj = QuestionModel.objects.get(question=question)
                #     print(question_obj)
                # except:
                #     print("Question model creating..")
                question_obj = QuestionModel.objects.get_or_create(question=question, created_by=request.user)
                for letter in letters:
                    answer = ws["{}{}".format(letter, i)].value
                    if answer and letter=="C":
                        answer_obj = AnswerModel.objects.get_or_create(answer=answer, question_id=question_obj[0].id, created_by=request.user, is_right=True)
                    elif answer:
                        answer_obj = AnswerModel.objects.get_or_create(answer=answer, question_id=question_obj[0].id, created_by=request.user)
                blank_limit = 4
            else:
                blank_limit -= 1
                if blank_limit == 0:
                    break
        except Exception as e:
            raise e
    return True

