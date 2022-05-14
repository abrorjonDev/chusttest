from openpyxl import load_workbook, Workbook
from .models import User, UserDocs


def names(full_name):
    first_name = None
    last_name = None
    father_name = None
    name = full_name.split(" ")
    if len(name)>3:
        first_name = name[1]
        last_name = name[0]
        father_name = name[2]+ ' '+ name[3]
    elif len(name)==3:
        first_name = name[1]
        last_name = name[0]
        father_name = name[2]
    else:
        first_name = name[1]
        last_name = name[0]
    return first_name, last_name, father_name

def create_students(file, created_by=None):
    wb = load_workbook(file)
    ws = wb.active

    for i in range(2, ws.max_row+1):
        username = ws.cell(i, 1).value #ID
        full_name = ws.cell(i, 2).value #full_name
        # birth_date = ws.cell(i, 3).value # birth_date
        # genre = ws.cell(i, 4).value #genre
        # nationality = ws.cell(i, 5).value
        # citizen = ws.cell(i, 6).value
        # state = ws.cell(i, 7).value
        # region = ws.cell(i, 8).value
        # pinfl = ws.cell(i, 9).value
        # doc_type = ws.cell(i, 10).value
        # serial = ws.cell(i, 11).value
        # doc_number = ws.cell(i, 12).value
        # who_gave = ws.cell(i, 13).value
        school = ws.cell(i, 14).value
        # state_organ = ws.cell(i, 15).value
        # region_organ = ws.cell(i, 16).value
        klass = ws.cell(i, 17).value
        modified, created = 0, 0
        try:
            student = User.objects.create(
                username=username, 
                first_name=names(full_name)[0], 
                last_name=names(full_name)[1],
                father_name=names(full_name)[2],
                school=school, 
                klass=klass,
                created_by=created_by
                )
            student.set_password("1")
            student.save()
            # student_doc, created = UserDocs.objects.get_or_create(
            #     user=student, 
            #     # birth_date=birth_date,
            #     # genre='a' if genre=='Женский' else 'e',
            #     # nationality=nationality,
            #     # citizen=citizen,
            #     # state=state,
            #     # region=region,
            #     # doc_type=doc_type,
            #     # pinfl=pinfl,
            #     # serial=serial,
            #     # doc_number=doc_number,
            #     # who_gave=who_gave,
            #     school=school,
            #     klass=klass,
            #     # state_organ=state_organ, 
            #     # region_organ=region_organ, 
            #     created_by=created_by
            # )
            created +=1

        except Exception as e:
            student = User.objects.get_or_create(username=username)[0]
            student.first_name=full_name.split(" ")[1]
            student.last_name=full_name.split(" ")[0]
            student.modified_by = created_by
            student.school = school
            student.klass = klass
            if not student.password:
                print("password saved")
                student.set_password('1')
            student.save()
            
            modified += 1
    return {'created':created, 'updated':modified, 'status':201}

