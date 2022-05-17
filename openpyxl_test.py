import openpyxl
import os
import datetime

file_data = openpyxl.load_workbook("./oQUVCHILAR_ROYHATI_3.xlsx")
ws = file_data.active()
print(datetime.datetime.now())
for i in range(ws.max_row()+1):
    print(ws.cell(i, 1).value)
    
print(datetime.datetime.now())
